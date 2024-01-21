#! /usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import io
import logging
import datetime
from openpyxl import load_workbook
from dataclasses import dataclass, field

# cSpell:ignore openpyxl datefmt levelname dataclass


@dataclass
class Report:
    file: str
    records: list = field(default_factory=list, init=False)

    def __post_init__(self):
        pass


@dataclass
class Record:
    row_number: int
    some_str: str
    some_int: int
    some_float: float
    some_date: datetime.date

    def __post_init__(self):
        pass


def parse(file, worksheet, row_of_label, callback=None):
    # iterate through rows of an Excel spreadsheet

    if callback is not None and not callable(callback):
        raise ValueError(f"callback given but it it not callable. It is a {type(callback)}.")

    wb = load_workbook(filename=file, read_only=True, data_only=True)
    ws = wb[worksheet]
    logging.debug(f'Slurp item information. Looking at worksheet "{ws.title}".')

    r = Report(file=file)

    logging.debug(f"Max row of this sheet is {ws.max_row}.")
    for row_number, row in enumerate(
        ws.iter_rows(min_row=row_of_label + 1, max_row=ws.max_row + 1, values_only=True), start=row_of_label + 1
    ):
        # In openpyxl, row and column are one based, not zero based.
        # Cell "A1" is row 1, col 1.
        # Note that when you iterate through rows, now a row is a tuple,
        # and it is zero based.
        # In this example, row[0] is column A.
        x = Record(
            row_number=row_number,
            some_str=row[0],
            some_int=row[1],
            some_float=row[2],
            some_date=row[3],
        )

        if callback is not None and callable(callback):
            callback(x)
        else:
            r.records.append(x)

    if callback is None:
        return r


def main():
    pass


if __name__ == "__main__":
    # https://qiita.com/jack-low/items/91bf9b5342965352cbeb
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    logger = logging.getLogger(__name__)
else:
    logger = logging.getLogger(f"__main__.{__name__}")
logger.setLevel(logging.DEBUG)

format_file = logging.Formatter("%(asctime)s %(filename)s: %(lineno)s: %(funcName)s - %(levelname)s: %(message)s")
file_handler = logging.FileHandler(str(sys.argv[0])[:-3] + ".log")
file_handler.setFormatter(format_file)
file_handler.setLevel(logging.DEBUG)
logger.addHandler(file_handler)

# add a Handler which writes INFO messages or higher to the console
format_console = logging.Formatter("%(filename)s: %(levelname)s %(funcName)s - %(message)s")
console_handler = logging.StreamHandler()
console_handler.setFormatter(format_console)
console_handler.setLevel(logging.INFO)
logger.addHandler(console_handler)

if __name__ == "__main__":
    main()
